import os
import random
import smtplib
from email.message import EmailMessage
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ==========================================
# CONFIGURAÇÕES GERAIS
# ==========================================
ARQUIVO_ENTRADA = os.getenv("ARQUIVO_ENTRADA", "produtos_atualizados.xlsx")
OUTPUT_DIR = Path(os.getenv("OUTPUT_DIR", "/data"))
ARQUIVO_SAIDA = OUTPUT_DIR / os.getenv("ARQUIVO_SAIDA", "precificacao_automatica.xlsx")

MARGEM_MINIMA_SEGURANCA = float(os.getenv("MARGEM_MINIMA_SEGURANCA", "15"))
TOLERANCIA_COMPETITIVA = float(os.getenv("TOLERANCIA_COMPETITIVA", "3"))
AJUSTE_URGENTE = float(os.getenv("AJUSTE_URGENTE", "8"))
ALVO_COMPETITIVO = float(os.getenv("ALVO_COMPETITIVO", "1.5"))

# envio opcional por e-mail
EMAIL_ENABLED = os.getenv("EMAIL_ENABLED", "false").strip().lower() == "true"
SMTP_HOST = os.getenv("SMTP_HOST", "")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASS = os.getenv("SMTP_PASS", "")
EMAIL_FROM = os.getenv("EMAIL_FROM", "")
EMAIL_TO = os.getenv("EMAIL_TO", "")
EMAIL_SUBJECT = os.getenv("EMAIL_SUBJECT", "Relatório automático de precificação TEC9")

# pasta persistente
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


# ==========================================
# FUNÇÕES DE NEGÓCIO
# ==========================================
def simular_mercado(preco: float) -> float:
    variacao = random.uniform(-0.10, 0.10)
    return round(preco * (1 + variacao), 2)


def classificar_status(dif_percentual: float) -> str:
    if dif_percentual > TOLERANCIA_COMPETITIVA:
        return "ACIMA DO MERCADO"
    elif dif_percentual < -TOLERANCIA_COMPETITIVA:
        return "ABAIXO DO MERCADO"
    return "COMPETITIVO"


def preco_minimo_seguro(custo: float, margem_minima: float) -> float:
    return round(custo * (1 + margem_minima / 100), 2)


def sugerir_preco(row: pd.Series) -> pd.Series:
    custo = float(row["CUSTO_TRATADO"])
    preco_venda = float(row["PRECO_VENDA"])
    preco_mercado = float(row["PRECO_MERCADO"])
    dif = float(row["DIFERENCA_%"])
    preco_min_seguro = float(row["PRECO_MINIMO_SEGURO"])

    preco_alvo = round(preco_mercado * (1 + ALVO_COMPETITIVO / 100), 2)
    preco_alvo = max(preco_alvo, preco_min_seguro)

    if dif > AJUSTE_URGENTE:
        acao = "AJUSTE URGENTE"
        preco_sugerido = preco_alvo
    elif dif > TOLERANCIA_COMPETITIVA:
        acao = "REDUZIR PRECO"
        preco_sugerido = preco_alvo
    elif dif < -AJUSTE_URGENTE:
        acao = "SUBIR PRECO"
        preco_sugerido = round(preco_mercado * 0.985, 2)
        preco_sugerido = max(preco_sugerido, preco_min_seguro)
    elif dif < -TOLERANCIA_COMPETITIVA:
        acao = "SUBIR PRECO"
        preco_sugerido = round((preco_venda + preco_mercado) / 2, 2)
        preco_sugerido = max(preco_sugerido, preco_min_seguro)
    else:
        acao = "MANTER"
        preco_sugerido = preco_venda

    if abs(preco_sugerido - preco_venda) < 0.01:
        preco_sugerido = preco_venda
        acao = "MANTER"

    return pd.Series([round(preco_sugerido, 2), acao])


def classificar_oportunidade(row: pd.Series) -> str:
    acao = row["ACAO_RECOMENDADA"]
    variacao = float(row["VARIACAO_SUGERIDA_%"])

    if acao == "AJUSTE URGENTE":
        return "PRIORIDADE MAXIMA"
    elif acao == "REDUZIR PRECO" and variacao < -3:
        return "CORRECAO IMPORTANTE"
    elif acao == "SUBIR PRECO" and variacao > 3:
        return "GANHO DE MARGEM"
    elif acao == "MANTER":
        return "ESTAVEL"
    return "AJUSTE MODERADO"


# ==========================================
# FUNÇÃO DE FORMATAÇÃO
# ==========================================
def formatar_planilha(ws):
    cor_cabecalho = PatternFill(fill_type="solid", fgColor="1F4E78")
    cor_branca = "FFFFFF"

    for cell in ws[1]:
        cell.fill = cor_cabecalho
        cell.font = Font(color=cor_branca, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            valor = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(valor))
        ws.column_dimensions[col_letter].width = max_length + 2

    ws.freeze_panes = "A2"


# ==========================================
# CARREGAMENTO E ANÁLISE
# ==========================================
def gerar_relatorio() -> dict:
    arquivo_entrada_path = Path(ARQUIVO_ENTRADA)

    if not arquivo_entrada_path.exists():
        raise FileNotFoundError(
            f"Arquivo de entrada não encontrado: {arquivo_entrada_path.resolve()}"
        )

    df = pd.read_excel(arquivo_entrada_path)

    if "PRODUTO" not in df.columns:
        df["PRODUTO"] = ""

    colunas_necessarias = ["SKU", "CUSTO_TRATADO", "PRECO_VENDA", "MARGEM_%"]
    for coluna in colunas_necessarias:
        if coluna not in df.columns:
            raise Exception(f"Coluna obrigatória não encontrada: {coluna}")

    df["CUSTO_TRATADO"] = pd.to_numeric(df["CUSTO_TRATADO"], errors="coerce")
    df["PRECO_VENDA"] = pd.to_numeric(df["PRECO_VENDA"], errors="coerce")
    df["MARGEM_%"] = pd.to_numeric(df["MARGEM_%"], errors="coerce")
    df = df.dropna(subset=["CUSTO_TRATADO", "PRECO_VENDA"]).copy()

    df["PRECO_MERCADO"] = df["PRECO_VENDA"].apply(simular_mercado)
    df["DIFERENCA_R$"] = (df["PRECO_VENDA"] - df["PRECO_MERCADO"]).round(2)
    df["DIFERENCA_%"] = (((df["PRECO_VENDA"] / df["PRECO_MERCADO"]) - 1) * 100).round(2)
    df["MARGEM_REAL_%"] = (((df["PRECO_VENDA"] - df["CUSTO_TRATADO"]) / df["PRECO_VENDA"]) * 100).round(2)
    df["STATUS"] = df["DIFERENCA_%"].apply(classificar_status)
    df["PRECO_MINIMO_SEGURO"] = df["CUSTO_TRATADO"].apply(
        lambda x: preco_minimo_seguro(float(x), MARGEM_MINIMA_SEGURANCA)
    )

    df[["PRECO_SUGERIDO", "ACAO_RECOMENDADA"]] = df.apply(sugerir_preco, axis=1)
    df["VARIACAO_SUGERIDA_R$"] = (df["PRECO_SUGERIDO"] - df["PRECO_VENDA"]).round(2)
    df["VARIACAO_SUGERIDA_%"] = (((df["PRECO_SUGERIDO"] / df["PRECO_VENDA"]) - 1) * 100).round(2)
    df["MARGEM_SUGERIDA_%"] = (((df["PRECO_SUGERIDO"] - df["CUSTO_TRATADO"]) / df["PRECO_SUGERIDO"]) * 100).round(2)
    df["OPORTUNIDADE"] = df.apply(classificar_oportunidade, axis=1)

    ordem_colunas = [
        "SKU",
        "PRODUTO",
        "CUSTO_TRATADO",
        "PRECO_VENDA",
        "MARGEM_%",
        "PRECO_MERCADO",
        "DIFERENCA_R$",
        "DIFERENCA_%",
        "MARGEM_REAL_%",
        "STATUS",
        "PRECO_MINIMO_SEGURO",
        "PRECO_SUGERIDO",
        "VARIACAO_SUGERIDA_R$",
        "VARIACAO_SUGERIDA_%",
        "MARGEM_SUGERIDA_%",
        "ACAO_RECOMENDADA",
        "OPORTUNIDADE",
    ]
    df = df[ordem_colunas]

    df_acima = df[df["STATUS"] == "ACIMA DO MERCADO"].copy().sort_values(by="DIFERENCA_%", ascending=False)
    df_abaixo = df[df["STATUS"] == "ABAIXO DO MERCADO"].copy().sort_values(by="DIFERENCA_%", ascending=True)
    df_oportunidade = df[
        df["OPORTUNIDADE"].isin(["PRIORIDADE MAXIMA", "CORRECAO IMPORTANTE", "GANHO DE MARGEM"])
    ].copy()

    mapa_prioridade = {
        "PRIORIDADE MAXIMA": 1,
        "CORRECAO IMPORTANTE": 2,
        "GANHO DE MARGEM": 3,
    }
    if not df_oportunidade.empty:
        df_oportunidade["PRIORIDADE_TMP"] = df_oportunidade["OPORTUNIDADE"].map(mapa_prioridade)
        df_oportunidade = df_oportunidade.sort_values(
            by=["PRIORIDADE_TMP", "VARIACAO_SUGERIDA_%"],
            ascending=[True, True],
        )
        df_oportunidade = df_oportunidade.drop(columns=["PRIORIDADE_TMP"])

    total_itens = len(df)
    qtd_competitivo = int((df["STATUS"] == "COMPETITIVO").sum())
    qtd_acima = int((df["STATUS"] == "ACIMA DO MERCADO").sum())
    qtd_abaixo = int((df["STATUS"] == "ABAIXO DO MERCADO").sum())
    qtd_ajuste_urgente = int((df["ACAO_RECOMENDADA"] == "AJUSTE URGENTE").sum())
    qtd_reduzir = int((df["ACAO_RECOMENDADA"] == "REDUZIR PRECO").sum())
    qtd_subir = int((df["ACAO_RECOMENDADA"] == "SUBIR PRECO").sum())
    qtd_manter = int((df["ACAO_RECOMENDADA"] == "MANTER").sum())

    resumo_df = pd.DataFrame({
        "INDICADOR": [
            "TOTAL DE PRODUTOS",
            "COMPETITIVO",
            "ACIMA DO MERCADO",
            "ABAIXO DO MERCADO",
            "AJUSTE URGENTE",
            "REDUZIR PRECO",
            "SUBIR PRECO",
            "MANTER",
        ],
        "QUANTIDADE": [
            total_itens,
            qtd_competitivo,
            qtd_acima,
            qtd_abaixo,
            qtd_ajuste_urgente,
            qtd_reduzir,
            qtd_subir,
            qtd_manter,
        ]
    })

    with pd.ExcelWriter(ARQUIVO_SAIDA, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Analise")
        resumo_df.to_excel(writer, index=False, sheet_name="Resumo")
        df_acima.to_excel(writer, index=False, sheet_name="Acima_Mercado")
        df_abaixo.to_excel(writer, index=False, sheet_name="Abaixo_Mercado")
        df_oportunidade.to_excel(writer, index=False, sheet_name="Oportunidade")

    wb = load_workbook(ARQUIVO_SAIDA)

    cor_verde = PatternFill(fill_type="solid", fgColor="C6EFCE")
    cor_vermelha = PatternFill(fill_type="solid", fgColor="FFC7CE")
    cor_amarela = PatternFill(fill_type="solid", fgColor="FFEB9C")
    cor_azul = PatternFill(fill_type="solid", fgColor="D9EAF7")
    cor_laranja = PatternFill(fill_type="solid", fgColor="FCE4D6")

    ws = wb["Analise"]
    formatar_planilha(ws)

    mapa_colunas = {}
    for i, cell in enumerate(ws[1], start=1):
        mapa_colunas[cell.value] = i

    colunas_moeda = [
        "CUSTO_TRATADO", "PRECO_VENDA", "PRECO_MERCADO",
        "DIFERENCA_R$", "PRECO_MINIMO_SEGURO",
        "PRECO_SUGERIDO", "VARIACAO_SUGERIDA_R$"
    ]
    colunas_percentual = [
        "MARGEM_%", "DIFERENCA_%", "MARGEM_REAL_%",
        "VARIACAO_SUGERIDA_%", "MARGEM_SUGERIDA_%"
    ]

    for row in range(2, ws.max_row + 1):
        for nome_coluna in colunas_moeda:
            ws.cell(row=row, column=mapa_colunas[nome_coluna]).number_format = 'R$ #,##0.00'
        for nome_coluna in colunas_percentual:
            ws.cell(row=row, column=mapa_colunas[nome_coluna]).number_format = '0.00'

        status_cell = ws.cell(row=row, column=mapa_colunas["STATUS"])
        if status_cell.value == "ACIMA DO MERCADO":
            status_cell.fill = cor_vermelha
            status_cell.font = Font(bold=True)
        elif status_cell.value == "ABAIXO DO MERCADO":
            status_cell.fill = cor_verde
            status_cell.font = Font(bold=True)
        elif status_cell.value == "COMPETITIVO":
            status_cell.fill = cor_amarela
            status_cell.font = Font(bold=True)

        acao_cell = ws.cell(row=row, column=mapa_colunas["ACAO_RECOMENDADA"])
        if acao_cell.value == "AJUSTE URGENTE":
            acao_cell.fill = cor_vermelha
            acao_cell.font = Font(bold=True)
        elif acao_cell.value == "REDUZIR PRECO":
            acao_cell.fill = cor_laranja
            acao_cell.font = Font(bold=True)
        elif acao_cell.value == "SUBIR PRECO":
            acao_cell.fill = cor_azul
            acao_cell.font = Font(bold=True)
        elif acao_cell.value == "MANTER":
            acao_cell.fill = cor_amarela

        oportunidade_cell = ws.cell(row=row, column=mapa_colunas["OPORTUNIDADE"])
        if oportunidade_cell.value == "PRIORIDADE MAXIMA":
            oportunidade_cell.fill = cor_vermelha
            oportunidade_cell.font = Font(bold=True)
        elif oportunidade_cell.value == "CORRECAO IMPORTANTE":
            oportunidade_cell.fill = cor_laranja
            oportunidade_cell.font = Font(bold=True)
        elif oportunidade_cell.value == "GANHO DE MARGEM":
            oportunidade_cell.fill = cor_azul
            oportunidade_cell.font = Font(bold=True)
        elif oportunidade_cell.value == "ESTAVEL":
            oportunidade_cell.fill = cor_amarela

    for nome_aba in ["Resumo", "Acima_Mercado", "Abaixo_Mercado", "Oportunidade"]:
        formatar_planilha(wb[nome_aba])

    ws_resumo = wb["Resumo"]
    for row in range(2, ws_resumo.max_row + 1):
        indicador = ws_resumo.cell(row=row, column=1).value
        if indicador == "COMPETITIVO":
            ws_resumo.cell(row=row, column=1).fill = cor_amarela
        elif indicador == "ACIMA DO MERCADO":
            ws_resumo.cell(row=row, column=1).fill = cor_vermelha
        elif indicador == "ABAIXO DO MERCADO":
            ws_resumo.cell(row=row, column=1).fill = cor_verde
        elif indicador == "AJUSTE URGENTE":
            ws_resumo.cell(row=row, column=1).fill = cor_vermelha
        elif indicador == "REDUZIR PRECO":
            ws_resumo.cell(row=row, column=1).fill = cor_laranja
        elif indicador == "SUBIR PRECO":
            ws_resumo.cell(row=row, column=1).fill = cor_azul
        elif indicador == "MANTER":
            ws_resumo.cell(row=row, column=1).fill = cor_amarela

    wb.save(ARQUIVO_SAIDA)

    resumo = {
        "total": total_itens,
        "competitivo": qtd_competitivo,
        "acima": qtd_acima,
        "abaixo": qtd_abaixo,
        "ajuste_urgente": qtd_ajuste_urgente,
        "reduzir_preco": qtd_reduzir,
        "subir_preco": qtd_subir,
        "manter": qtd_manter,
        "arquivo": str(ARQUIVO_SAIDA),
    }
    return resumo


# ==========================================
# ENVIO DE E-MAIL OPCIONAL
# ==========================================
def enviar_email_resumo(resumo: dict):
    if not EMAIL_ENABLED:
        print("\n📨 Envio por e-mail desativado.")
        return

    campos_obrigatorios = [SMTP_HOST, SMTP_USER, SMTP_PASS, EMAIL_FROM, EMAIL_TO]
    if any(not campo for campo in campos_obrigatorios):
        print("\n⚠️ E-mail não enviado: variáveis SMTP incompletas.")
        return

    corpo = f"""
Olá,

Segue o resumo automático da precificação TEC9:

Total de produtos: {resumo['total']}
Competitivo: {resumo['competitivo']}
Acima do mercado: {resumo['acima']}
Abaixo do mercado: {resumo['abaixo']}
Ajuste urgente: {resumo['ajuste_urgente']}
Reduzir preço: {resumo['reduzir_preco']}
Subir preço: {resumo['subir_preco']}
Manter: {resumo['manter']}

Arquivo gerado:
{resumo['arquivo']}

Mensagem automática enviada pelo Railway.
""".strip()

    msg = EmailMessage()
    msg["Subject"] = EMAIL_SUBJECT
    msg["From"] = EMAIL_FROM
    msg["To"] = EMAIL_TO
    msg.set_content(corpo)

    arquivo_path = Path(resumo["arquivo"])
    with open(arquivo_path, "rb") as f:
        conteudo = f.read()

    msg.add_attachment(
        conteudo,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=arquivo_path.name,
    )

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
        server.starttls()
        server.login(SMTP_USER, SMTP_PASS)
        server.send_message(msg)

    print("\n✅ E-mail enviado com sucesso.")


# ==========================================
# MAIN
# ==========================================
if __name__ == "__main__":
    resumo = gerar_relatorio()

    print("\n📊 RESUMO:")
    print(f"TOTAL DE PRODUTOS: {resumo['total']}")
    print(f"COMPETITIVO: {resumo['competitivo']}")
    print(f"ACIMA DO MERCADO: {resumo['acima']}")
    print(f"ABAIXO DO MERCADO: {resumo['abaixo']}")
    print(f"AJUSTE URGENTE: {resumo['ajuste_urgente']}")
    print(f"REDUZIR PRECO: {resumo['reduzir_preco']}")
    print(f"SUBIR PRECO: {resumo['subir_preco']}")
    print(f"MANTER: {resumo['manter']}")

    print("\n✅ Precificação automática concluída com sucesso!")
    print(f"📁 Arquivo gerado: {resumo['arquivo']}")

    enviar_email_resumo(resumo)
